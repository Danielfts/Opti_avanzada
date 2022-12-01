import pandas as pd
import gurobipy as gp 
from gurobipy import GRB
import pandas as pd
import numpy as np 
import pyflakes
from openpyxl import*


book = load_workbook("/Users/spheal/Desktop/a.xlsx")
sheet = book.active 
datos = book.active

demanda = [] #demanda
for fila in range(1,24):
        demanda.append(sheet.cell(  12,fila+5).value)

miuk= [] #mius
for fila in range(1,24):
        miuk.append(sheet.cell(  14,fila+5).value)
columnas= []
matriz=[]
for i in range(1,9):
    fila= []

    for j in range(1,25):
        fila.append(sheet.cell(i+18,j+5).value)
    matriz.append(fila)
        
matriz =np.array(matriz)
dfm=pd.DataFrame(matriz)


demanda = np.array(demanda)


    
